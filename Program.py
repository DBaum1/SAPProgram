#!/usr/bin/python
#!python3
from tkinter import *
from tkinter.filedialog import askopenfilename
from tkinter import messagebox
import openpyxl
from openpyxl.utils import cell
import os
import time
import pywinauto
import re
import win32api
import pyautogui
from pywinauto.application import Application
from pywinauto import keyboard
from pywinauto.timings import Timings
from shutil import copyfile
from configparser import ConfigParser, NoSectionError, DuplicateSectionError,\
                     NoOptionError, MissingSectionHeaderError

#constants
COLOR = '#063256'
FONT = 11
DEFAULT_SIZE = '560x450'
MIN_WIDTH = 550
MIN_HEIGHT = 440
PATH = 'C:\Program Files (x86)\SAP\FrontEnd\SAPgui\saplogon.exe'
#Listings to check SAP for
LISTINGS = [('Contract Number','1'), ('OA Amount', '6'), ('OA Net', '7'),
            ('Validity Start Date', '10'), ('Expiration Date', '11')]      
ENTRY_LIST = [None] * len(LISTINGS) #Store references to grid entries
FIELD_LENGTH = 70 #length of SAP field

#copies original file, timestamps backup. Returns path of backup
def save_backup(file_path):
    src = file_path
    components = os.path.splitext(src)
    root = components[0]
    ext = components[1]
    time_tuple = time.localtime()
    format_time = time.strftime('_%m_%d_%Y_%Hh_%Mm_%Ss', time_tuple)
    root_format = root + format_time
    dest = root_format + ext
    copyfile(src, dest)
    return dest
    
#Creates config file
def init_config():
    config = ConfigParser()
    config.write('config.ini')
    config.add_section('main')
    config.set('main', 'path', PATH)
    with open('config.ini', 'w') as f:
        for i in range(len(LISTINGS)):
            curr_tuple = LISTINGS[i]
            config.set('main', curr_tuple[0], curr_tuple[1])
        config.write(f)
    f.close()

#Reads from config
def read_from_config():
    config = ConfigParser()
    try:
        config.read('config.ini')
        #loop through entry fields
        for i in range(len(ENTRY_LIST)):
            curr_entry = ENTRY_LIST[i]
            try:
                val = config.get('main', LISTINGS[i][0])
                new_tuple = (LISTINGS[i][0], str(val))                     
                LISTINGS[i] = new_tuple
                #Convert from int to alphanumeric column
                val = cell.get_column_letter(int(val))
                curr_entry.insert(0, str(val))
            #Config corrupted, remake
            except (NoSectionError, DuplicateSectionError,
                    MissingSectionHeaderError):
                init_config()
                read_from_config()
            except NoOptionError:
                pass
            except ValueError:
                new_tuple = (LISTINGS[i][0], '-1')
                LISTINGS[i] = new_tuple
                curr_entry.insert(0, '')
    #Config corrupted, remake
    except (MissingSectionHeaderError):
        init_config()
        read_from_config()

#Write user entered data to config file
def write_to_config(btn):    
    config = ConfigParser()
    config.read('config.ini')
    for i in range(len(ENTRY_LIST)):
        #loop through entry fields
        val = ENTRY_LIST[i].get()
        try:
            #Convert alphanumeric column to int
            val = str(cell.column_index_from_string(val))
        except ValueError:
            val = '-1'
            new_tuple = (LISTINGS[i][0], '-1')
            LISTINGS[i] = new_tuple
        #Update LISTINGS
        new_tuple = (LISTINGS[i][0], val)                  
        LISTINGS[i] = new_tuple
        try:
            config.set('main', LISTINGS[i][0], val)
        except (NoSectionError, DuplicateSectionError,
            MissingSectionHeaderError):
            init_config()
            config.read('config.ini')
            config.set('main', LISTINGS[i][0], val)
        except NoOptionError:
            pass
    # save to config file
    with open('config.ini', 'w') as configfile:
        config.write(configfile)
    btn.configure(state=NORMAL)
    
class SAPTransferGUI:

    def __init__(self, master):
        self.master = master
        self.initialize()
        
    #Initializes main GUI components
    def initialize(self):
        #Labels
        label1_text = ("Program to transfer data from SAP database to Excel"
        " spreadsheet. Please log into SAP and navigate to the agreement"
        " number page before using.\n")
        label1 = Label(text=label1_text, wraplength=500, font=(None, FONT),
                      foreground='white', background=COLOR).pack(fill=X)
        label2_text = (" NOTE: This program creates a timestamped backup of"
                " any file it modifies. It is HIGHLY recommended NOT to"
                " delete the backup until you have verified all the new"
               " information is valid.\n")
        label2 = Label(text=label2_text, wraplength=500, font=(None, FONT),
                    foreground='red', background='white').pack(fill=X)
        label3_text = ("WARNING: Terminating the program before it has"
               " finished the data transfer will cause all changes to be"
               " rolled back, necessitating restarting to continue. Press"
               " CTRL-ALT-DELETE to stop GUI automation.")
        label3 = Label(text=label3_text, wraplength=500, font=(None, FONT),
                foreground='white', background=COLOR).pack(fill=X)
        
        #File Chooser Button
        file_button = Button(text='Select Excel file',font=(None, FONT),
                command=lambda: self.show_file_chooser(path_entry, import_btn))
        file_button.pack(pady=10)

        #Entry to show user the destination path they chose
        dest = StringVar()
        path_entry = Entry(textvariable=dest, width=50, justify=LEFT,
                      state='readonly',font=(None, FONT))
        path_entry.pack(pady=5)

        #Sheet
        sheet_label = Label(text="Enter Excel sheet name ex: Services "
                            + "(case sensitive)", font=(None, FONT))
        sheet_label.pack(pady=5)
        dest = StringVar()
        dest.set('Chemicals')
        sheet_entry = Entry(font=(None, FONT), textvariable=dest)
        sheet_entry.pack(pady=10)

        #Column Info
        col_info_btn = Button(root, text="Enter/Verify Column Information",
            font=(None, FONT), command=lambda:
            (col_info_btn.config(state=DISABLED), ColTable(self, col_info_btn)))
        col_info_btn.pack(pady=10)

        #Import Button, initially disabled
        import_btn = Button(root, text="Import from SAP", state=DISABLED,
          command=lambda:(import_btn.config(state=DISABLED),
          self.import_data(path_entry, sheet_entry, import_btn)),
                            font=(None, FONT))
        import_btn.pack()                                               

    #Excel file selection dialog
    def show_file_chooser(self, path_entry, btn):
        path_entry.configure(state='normal')
        file_path = askopenfilename(parent=None, title = "Select file",
                        filetypes = [(("Excel (.xlsx)","*.xlsx"))])
        #show user the filepath they selected
        if file_path is not None and len(file_path) > 0:
            path_entry.delete(0, END)
            path_entry.insert(0, file_path)
            path_entry.configure(state='readonly')
            btn.config(state=NORMAL)

    #Imports data from SAP to Excel doc
    def import_data(self, path_entry, sheet_entry, btn):
        file_path = path_entry.get()
        try:
            wb = openpyxl.load_workbook(file_path)
            #get sheet user entered
            sheetname = sheet_entry.get()
            sheets = wb.sheetnames
            sheet = wb[sheetname]
            max_row = sheet.max_row
            max_col = sheet.max_column
            start_row = self.get_start_row(sheet, max_row)
            backup_path = save_backup(file_path)
            #transfer started
            try:
                app = Application(backend='uia').connect(path=PATH)
                #Display Contract:Initial Screen
                dlg_spec = app.Display_Contract_Initial_Screen
                actionable_dlg = dlg_spec.wait('visible')
                rect = dlg_spec['Pane6'].rectangle()
                x_ref = rect.left
                y_ref = rect.top
                pyautogui.moveTo(x_ref, y_ref)
                pyautogui.click()
                #Start transfer
                for r in range(start_row, max_row):
                    #index of LISTINGS 'Contract Number' column
                    contract_col = int(LISTINGS[0][1])
                    contract_num = sheet.cell(row=r, column=contract_col).value                        
                    #only try to transfer data if there's a contract number
                    #to search
                    if self.is_contract_num(contract_num):
                        pyautogui.typewrite(str(contract_num), interval=0.25)
                        pyautogui.press('enter', interval=5)
                        #Header details
                        pyautogui.press('f6', interval=5)
                        self.sap_transfer(sheet, r, contract_num, app)
                wb.save(file_path)
                #App done with transfer
                messagebox.showinfo("Finished!", "Transfer finished.")

            except pywinauto.application.ProcessNotFoundError:
                text = ("Please make sure that SAP is running and"
                        " you have navigated to Display Contract:Initial"
                        " Screen title page. If the contract"
                        " agreement page is open but you are"
                        " still getting this error, you will have"
                        " to change the PATH variable in the"
                        " config.ini file to the path of the"
                        " SAPLogon.exe. Then restart the import.")
                messagebox.showerror("Program not found!", text)
                #restores original file (dest) from backup (src) in case of
                #premature exit
                copyfile(backup_path, file_path)
                #Backup no longer needed
                os.remove(backup_path)
            except pywinauto.timings.TimeoutError:
                text = ("Sytem took too long. Please restart this program.")
                messagebox.showerror("Program not found!", text)
        #File no longer exists at path
        except IOError:
            text = ("File not found at selected path. Check to make sure it"
                    " wasn't deleted or moved.")
            messagebox.showerror("File not found!", text)
        except KeyError:
            text = ("Check sheet entry field for spelling, spacing, and"
            " capitalization. It must exactly match the Excel doc sheet name.")
            messagebox.showerror("Sheet not found!", text)
        btn.config(state=NORMAL)

    #transfers data from SAP fields to excel file
    def sap_transfer(self, sheet, row, contract_num, app):
        Timings.Slow()
        #Display Contract:Header Data
        dlg_spec = app.Display_Contract_Header_Data
        actionable_dlg = dlg_spec.wait('visible')
        rect = app.Display_Contract_Header_Data['Pane6'].rectangle()
        #Get coordinates that will be used as reference to get the
        #data from SAP fields
        x_ref = rect.left
        y_ref = rect.top
        #get agreement start
        valid_start = None
        x = x_ref + 125
        y = y_ref + 125
        pyautogui.moveTo(x, y)
        pyautogui.click()
        pyautogui.click()
        pywinauto.mouse.press(button='left', coords=(x+FIELD_LENGTH-1, y))
        time.sleep(0.2)
        valid_start = pyautogui.hotkey('ctrl', 'c')
        time.sleep(0.2)
        if valid_start is not None:
            sheet.cell(row=row, column=LISTINGS[3][1], value=valid_start)
        pywinauto.mouse.release(button='left', coords=(x+FIELD_LENGTH, y))
        #get agreement end
        valid_end = None
        x = x_ref + 330
        y = y_ref + 125
        pyautogui.moveTo(x, y)
        pyautogui.click()
        pyautogui.click()
        pywinauto.mouse.press(button='left', coords=(x+FIELD_LENGTH-1, y))
        time.sleep(0.2)
        valid_end = pyautogui.hotkey('ctrl', 'c')
        time.sleep(0.2)
        if valid_end is not None:
            sheet.cell(row=row, column=LISTINGS[4][1], value=valid_end)
        pywinauto.mouse.release(button='left', coords=(x+FIELD_LENGTH, y))
        #get OA net
        oa_net = None
        x = x_ref + 161
        y = y_ref + 470
        pyautogui.moveTo(x, y)
        pyautogui.click()
        pyautogui.click()
        pywinauto.mouse.press(button='left', coords=(x+102, y))
        time.sleep(0.2)
        oa_net = pyautogui.hotkey('ctrl', 'c')
        time.sleep(0.2)
        if oa_net is not None:
            sheet.cell(row=row, column=LISTINGS[2][1], value=oa_net)
        pywinauto.mouse.release(button='left', coords=(x+101, y))
        #return to contract agreement page
        pyautogui.press('f3', interval=3)
        
    #determines if data is a contract number (10 digits)
    def is_contract_num(self, val):
        #format of contract agreement numbers
        num_format = re.compile('[0-9]{10}')
        #index of LISTINGS 'Contract Number' column
        found = num_format.match(str(val))
        return found
    
    #get first row of contract data
    def get_start_row(self, sheet, max_row):
        contract_col = int(LISTINGS[0][1])
        try:
            for r in range(1, max_row):
                curr_cell = sheet.cell(row=r, column=contract_col)
                val = curr_cell.value
                #found first cell with contract number
                if self.is_contract_num(val):
                    return curr_cell.row
        except ValueError:
            text = ("No contract agreement numbers found in sheet!\n"
            " Please check to make sure that you have filled in the"
            " contract agreement (column info) field with a column that exists"
            " on the sheet. Else, check that the sheet contains valid"
            " agreement numbers. Valid input is defined as a sequence of 10"
            " numbers, 0-9 with no letters. \nEx: 4000954312")
            messagebox.showerror("Invalid entry!", text)
            
#Table where user inputs what information is in each column
class ColTable:
    def __init__(self, master, btn):
        self.master = master
        self.btn = btn
        self.initialize()
        
    #Initializes column table components
    def initialize(self):
        table = Toplevel(pady=10, padx=10)
        table.title("Set Column Info")
        table.geometry(DEFAULT_SIZE)
        label_text = ("Enter the column letter that the relevant information"
                      " can be found in (ex: A, B, AA etc. - Must be"
                      " capitalized). If a column's value is calculated with"
                      " a function, leave that field blank.")
        col_table_label = Label(table, text=label_text, wraplength = 500,
                                font=(None, FONT))
        col_table_label.grid(columnspan=len(LISTINGS))
        self.fill_grid(table)
        table.protocol("WM_DELETE_WINDOW", lambda:(write_to_config(self.btn),
                                               table.destroy()))
        
    #Fills entries with values from config file
    def fill_grid(self, parent):
        num_rows = len(LISTINGS)
        for i in range(num_rows):
            #initialize labels
            curr_tuple = LISTINGS[i]
            curr_label = Label(parent, text=curr_tuple[0], font=(None, FONT))
            curr_label.grid(row=i+1, column=2, padx=5,pady=5)
            #initialize entry fields
            curr_entry = Entry(parent, font=(None, FONT))
            curr_entry.grid(row=i+1, column=3, padx=5,pady=5)
            ENTRY_LIST[i] = curr_entry
        #config file exists
        try:
            f = open('config.ini', 'r')
        #create config from scratch
        except IOError:
            init_config()
        #read in values from config file
        read_from_config()

#App
root = Tk()
root.title("SAP to Excel")
root.configure(background=COLOR, pady=10, padx=10)
root.geometry(DEFAULT_SIZE)
root.minsize(width=MIN_WIDTH, height = MIN_HEIGHT)
gui = SAPTransferGUI(root)
root.mainloop()
